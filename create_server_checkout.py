import shutil
from openpyxl import load_workbook
from datetime import datetime, timedelta
 
############################################################################
######### This file creates a new server checkout for VINNY ABQ ############
############################################################################
 

def create_checkout(start_date = None):
        print("\nThis script generates a new server checkout .xlsx.\n"
        "It will be named MM.DD-MM.YY in a two week increment.\n"
        "It will be located in the destination path.")
 
        # Get start date and end date
        start_date = input("Enter the start date (MM.DD): ")
 
        # Start_date = '06.08'
        start_date = datetime.strptime(start_date, "%m.%d")
        end_date = start_date + timedelta(days = 13)
 
        # Make the datetime objects strings
        start_date_str = datetime.strftime(start_date, "%m.%d")
        end_date_str = datetime.strftime(end_date, "%m.%d")
 
        # Create the new file name (MM.DD-MM.DD)
        new_file_name = f'{start_date_str}-{end_date_str}'
 
        # Source and destination paths
        source_file = 'C:\\Users\\info\\OneDrive\\Desktop\\Matthew\\Python\\CreateServerCheckout\\ServerCheckoutTemplate.xlsx'  
        destination_path = f'C:\\Users\\info\\OneDrive\\Desktop\\Matthew\\Python\\CreateServerCheckout\\{new_file_name}.xlsx'
        
        # Copy and rename the file
        shutil.copy(source_file, destination_path)
        print(f'\nSuccessfully created new server checkout for {start_date_str} - {end_date_str}')
        print(f'The new file is located at {destination_path}\n')      
 
 
        # Open the workbook
        wb = load_workbook(destination_path)
 
        # Remove "NO EXPO"
        ws = wb["NO EXPO"]
        wb.remove(ws)
 
        # Rename EXPO to start_date_str
        ws = wb['EXPO']
        ws.title = start_date_str
 
        # Create 14 copies of "EXPO" and give appropriate dates as names
        for i in range(13):
                new_sheet_date = start_date + timedelta(days= i+1)
                new_sheet_name = datetime.strftime(new_sheet_date, "%m.%d")
                new_sheet = wb.copy_worksheet(ws)
                new_sheet.title = new_sheet_name
                wb.worksheets.insert(1 + i, new_sheet)
 

        # Move 'Summary of Employee Tips' to the end of the file
        sheet_to_move = wb['Summary of Employee Tips']
        wb.move_sheet(sheet_to_move, offset=len(wb.sheetnames)-1)
 
        # Save the workbook
        wb.save(destination_path)
       
create_checkout()
